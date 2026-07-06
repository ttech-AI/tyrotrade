#!/usr/bin/env python3
"""Generate src/data/powerbiPL.ts from the Power BI Excel exports.

Static "LIVE REALIZED – PROJECTED P&L" snapshots, one per financial year,
rendered by the "Power BI Version" table on the E.M Bakış dashboard as a fixed
reference. The table for a given FY is shown ONLY when that FY is the selected
filter (a 25-26 export is meaningless under a 24-25 filter and vice-versa).

Auto-discovers every year pair sitting in the tyro-project-mcp repo root
(= parent of tyrotrade-repo):
  - "Live Realized - Projected P&L <FY>.xlsx"                 → monthly rows
  - "Live Realized - Projected P&L Detailed Matrix <FY>.xlsx" → segment drill-down
where <FY> is "YY-YY" (e.g. "24-25", "25-26"). Drop in a new pair + re-run:

    python scripts/build-powerbi-pl.py
"""
import os
import re
import sys
import io
import glob
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(SCRIPT_DIR)                 # tyrotrade-repo
OUTER = os.path.dirname(REPO)                      # tyro-project-mcp-dev (Excels live here)
OUT_TS = os.path.join(REPO, "src", "data", "powerbiPL.ts")

# "Live Realized - Projected P&L 25-26.xlsx" → FY "25-26"; the Detailed Matrix
# file has "Detailed Matrix " before the FY so it won't match this pattern.
MAIN_RE = re.compile(r"^Live Realized - Projected P&L (\d{2}-\d{2})\.xlsx$")

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def month_key(label):
    """'Jul-25' -> '2025-07'. Returns None for non-month labels (Total, blank)."""
    if not label or not isinstance(label, str) or "-" not in label:
        return None
    mon, yy = label.strip().split("-", 1)
    m = _MONTHS.get(mon.strip().lower()[:3])
    if m is None or not yy.strip().isdigit():
        return None
    return f"20{int(yy):02d}-{m:02d}"


def num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        mk = month_key(r[0])
        if mk is None:
            continue  # skips Total + blank rows
        rows.append({
            "monthKey": mk,
            "projQtyTons": num(r[1]),
            "projRevenueUsd": num(r[2]),
            "projPLUsd": num(r[3]),
            "budgetUsd": num(r[4]),
            "realQtyTons": num(r[5]),
            "realRevenueUsd": num(r[6]),
            "realPLUsd": num(r[7]),
        })
    wb.close()
    return rows


def load_detail(path):
    """monthKey -> list of {segment, projPL, realPL, budget}. Header layout:
    row1 = month labels (col 1, 4, 7, …), row2 = [Projected P&L, Live Realized
    P&L, Projected Budget] per month, col0 = Segment Name."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    grid = list(ws.iter_rows(values_only=True))
    wb.close()
    header = grid[0]
    idx_to_key = {}
    for i in range(12):
        col = 1 + 3 * i
        if col < len(header):
            idx_to_key[i] = month_key(header[col])
    out = {mk: [] for mk in idx_to_key.values() if mk}
    for row in grid[2:]:
        seg = (row[0] or "").strip() if isinstance(row[0], str) else ""
        if not seg or seg.lower() in ("total", "grand total", "toplam"):
            continue
        for i in range(12):
            mk = idx_to_key.get(i)
            if not mk:
                continue
            proj = num(row[1 + 3 * i]) if 1 + 3 * i < len(row) else 0.0
            real = num(row[2 + 3 * i]) if 2 + 3 * i < len(row) else 0.0
            bud = num(row[3 + 3 * i]) if 3 + 3 * i < len(row) else 0.0
            if proj == 0.0 and real == 0.0 and bud == 0.0:
                continue
            out[mk].append({
                "segment": seg, "projPLUsd": proj, "realPLUsd": real, "budgetUsd": bud,
            })
    return out


def discover():
    """Returns sorted list of (fy, main_path, detail_path) tuples."""
    found = []
    for path in glob.glob(os.path.join(OUTER, "*.xlsx")):
        m = MAIN_RE.match(os.path.basename(path))
        if not m:
            continue
        fy = m.group(1)
        detail = os.path.join(OUTER, f"Live Realized - Projected P&L Detailed Matrix {fy}.xlsx")
        if not os.path.exists(detail):
            print(f"  ! {fy}: main found but no Detailed Matrix — skipping")
            continue
        found.append((fy, path, detail))
    return sorted(found, key=lambda x: x[0])


def fmt(v):
    return str(int(v)) if v == int(v) else repr(round(v, 6))


def ts_str(s):
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'


def main():
    years = discover()
    if not years:
        print("No 'Live Realized - Projected P&L <FY>.xlsx' files found in", OUTER)
        sys.exit(1)

    data = {}  # fy -> (rows, detail)
    for fy, main_path, detail_path in years:
        rows = load_main(main_path)
        detail = load_detail(detail_path)
        data[fy] = (rows, detail)
        tp = sum(m["realPLUsd"] for m in rows)
        tb = sum(m["budgetUsd"] for m in rows)
        print(f"FY {fy}: {len(rows)} months, Σ realized P&L={tp:,.0f}, Σ budget={tb:,.0f}")

    L = []
    L.append("// AUTO-GENERATED by scripts/build-powerbi-pl.py — DO NOT EDIT BY HAND.")
    L.append("// Static Power BI export snapshots (one per financial year) for the")
    L.append("// \"Power BI Version\" reference table on the E.M Bakış dashboard. Re-run the")
    L.append("// script after adding/replacing a \"Live Realized - Projected P&L … .xlsx\" pair.")
    L.append("")
    L.append("export interface PowerBIPLMonthRow {")
    L.append("  /** 'YYYY-MM' of the FY month. */")
    L.append("  monthKey: string;")
    L.append("  projQtyTons: number;")
    L.append("  projRevenueUsd: number;")
    L.append("  projPLUsd: number;")
    L.append("  budgetUsd: number;")
    L.append("  realQtyTons: number;")
    L.append("  realRevenueUsd: number;")
    L.append("  realPLUsd: number;")
    L.append("}")
    L.append("")
    L.append("export interface PowerBIPLSegmentRow {")
    L.append("  segment: string;")
    L.append("  projPLUsd: number;")
    L.append("  realPLUsd: number;")
    L.append("  budgetUsd: number;")
    L.append("}")
    L.append("")
    L.append("export interface PowerBIPLYear {")
    L.append("  /** FinancialYear.label this snapshot belongs to (e.g. '25-26'). */")
    L.append("  fy: string;")
    L.append("  /** 12 FY-month rows, Jul → Jun. */")
    L.append("  rows: PowerBIPLMonthRow[];")
    L.append("  /** monthKey → per-segment breakdown (drill-down). */")
    L.append("  segments: Record<string, PowerBIPLSegmentRow[]>;")
    L.append("}")
    L.append("")
    L.append("/** Every FY export we have, keyed by FinancialYear.label. The dashboard")
    L.append(" *  renders the Power BI Version table only for a selected FY that is a key")
    L.append(" *  here. */")
    L.append("export const POWERBI_PL_BY_FY: Record<string, PowerBIPLYear> = {")
    for fy, _, _ in years:
        rows, detail = data[fy]
        L.append(f"  {ts_str(fy)}: {{")
        L.append(f"    fy: {ts_str(fy)},")
        L.append("    rows: [")
        for m in rows:
            L.append(
                "      { monthKey: \"%s\", projQtyTons: %s, projRevenueUsd: %s, projPLUsd: %s, "
                "budgetUsd: %s, realQtyTons: %s, realRevenueUsd: %s, realPLUsd: %s }," % (
                    m["monthKey"], fmt(m["projQtyTons"]), fmt(m["projRevenueUsd"]),
                    fmt(m["projPLUsd"]), fmt(m["budgetUsd"]), fmt(m["realQtyTons"]),
                    fmt(m["realRevenueUsd"]), fmt(m["realPLUsd"]),
                )
            )
        L.append("    ],")
        L.append("    segments: {")
        for m in rows:
            mk = m["monthKey"]
            L.append(f"      \"{mk}\": [")
            for s in detail.get(mk, []):
                L.append(
                    "        { segment: %s, projPLUsd: %s, realPLUsd: %s, budgetUsd: %s }," % (
                        ts_str(s["segment"]), fmt(s["projPLUsd"]), fmt(s["realPLUsd"]), fmt(s["budgetUsd"]),
                    )
                )
            L.append("      ],")
        L.append("    },")
        L.append("  },")
    L.append("};")
    L.append("")

    os.makedirs(os.path.dirname(OUT_TS), exist_ok=True)
    with open(OUT_TS, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(L))
    print(f"\nwrote {OUT_TS}  ({', '.join(fy for fy, _, _ in years)})")


if __name__ == "__main__":
    main()
